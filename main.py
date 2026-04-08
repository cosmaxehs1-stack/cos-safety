import os
import io
import json
import hashlib
import re
import secrets
import uuid
import zipfile
import base64
from collections import Counter
from datetime import datetime, date, timedelta
from typing import Optional, Union
from xml.etree import ElementTree as ET

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request, Response
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
import psycopg2
import psycopg2.extras
import psycopg2.pool

app = FastAPI(title="COS-Safety Dashboard")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = os.environ.get("DATA_DIR", "uploads")
IMAGE_DIR = os.path.join(UPLOAD_DIR, "images")
DATA_FILE = os.path.join(UPLOAD_DIR, "current_data.json")
PASSWORD = os.environ.get("DASHBOARD_PASSWORD", "2026")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin2026")
DATABASE_URL = os.environ.get("DATABASE_URL")
SESSION_TOKENS: set[str] = set()
ADMIN_TOKENS: set[str] = set()

# --- Data Cache ---
_data_cache: list[dict] | None = None
_data_cache_time: float = 0
_DATA_CACHE_TTL = 2.0  # seconds

import time as _time

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(IMAGE_DIR, exist_ok=True)


# --- Database ---
_db_pool = None

def _get_pool():
    global _db_pool
    if _db_pool is None and DATABASE_URL:
        _db_pool = psycopg2.pool.ThreadedConnectionPool(1, 10, DATABASE_URL, connect_timeout=10)
    return _db_pool

def get_db(retries=3):
    pool = _get_pool()
    if pool:
        for attempt in range(retries):
            try:
                conn = pool.getconn()
                conn.autocommit = False
                return conn
            except Exception as e:
                print(f"[DB] 연결 실패 (시도 {attempt+1}/{retries}): {e}")
                if attempt == retries - 1:
                    raise
    for attempt in range(retries):
        try:
            conn = psycopg2.connect(DATABASE_URL, connect_timeout=10)
            return conn
        except Exception as e:
            print(f"[DB] 연결 실패 (시도 {attempt+1}/{retries}): {e}")
            if attempt == retries - 1:
                raise

def release_db(conn):
    pool = _get_pool()
    if pool:
        try:
            pool.putconn(conn)
        except Exception:
            try: conn.close()
            except: pass
    else:
        try: conn.close()
        except: pass


def init_db():
    if not DATABASE_URL:
        return
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS records (
            _id TEXT PRIMARY KEY,
            data JSONB NOT NULL
        )
    """)
    cur.execute("DROP TABLE IF EXISTS excel_files")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS weekly_snapshots (
            id TEXT PRIMARY KEY,
            year INT NOT NULL,
            quarter INT NOT NULL,
            month INT NOT NULL,
            week INT NOT NULL,
            saved_at TEXT NOT NULL,
            data JSONB NOT NULL
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS executive_comments (
            id TEXT PRIMARY KEY,
            role TEXT NOT NULL,
            content TEXT NOT NULL,
            week_key TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT
        )
    """)
    try:
        cur.execute("ALTER TABLE executive_comments ADD COLUMN week_key TEXT NOT NULL DEFAULT ''")
    except:
        conn.rollback()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS executive_notifications (
            id TEXT PRIMARY KEY,
            comment_id TEXT NOT NULL,
            message TEXT NOT NULL,
            created_at TEXT NOT NULL,
            is_read BOOLEAN DEFAULT FALSE
        )
    """)
    conn.commit()
    cur.close()
    release_db(conn)


@app.on_event("startup")
def on_startup():
    print(f"[DB] DATABASE_URL set: {bool(DATABASE_URL)}")
    try:
        init_db()
        print("[DB] init_db completed successfully")
    except Exception as e:
        print(f"[DB] init_db error: {e}")


# --- Auth ---
@app.post("/api/login")
async def login(request: Request):
    return {"token": "public"}


@app.post("/api/admin/login")
async def admin_login(request: Request):
    body = await request.json()
    pw = body.get("password", "")
    if pw != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="관리자 비밀번호가 올바르지 않습니다.")
    token = secrets.token_hex(16)
    ADMIN_TOKENS.add(token)
    return {"admin_token": token}


def verify_token(request: Request):
    pass


def verify_admin(request: Request):
    token = request.headers.get("X-Admin-Token", "")
    if token not in ADMIN_TOKENS:
        raise HTTPException(status_code=403, detail="관리자 권한이 필요합니다.")


# --- Excel Parsing ---
def parse_date(val) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s.split(" ")[0] if " " in s and "." not in s else s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def parse_number(val) -> int:
    if val is None:
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0


def extract_location_group(location: str) -> str:
    """소분류 장소 그룹을 반환"""
    if not location:
        return "기타(전공장)"
    loc = location.strip()
    if "화성" in loc:
        for i in [1, 2, 3, 5]:
            if f"{i}공장" in loc or f"{i} 공장" in loc:
                return f"화성{i}공장"
        return "기타(전공장)"
    if "평택" in loc:
        for i in [1, 2]:
            if f"{i}공장" in loc or f"{i} 공장" in loc:
                return f"평택{i}공장"
        return "기타(전공장)"
    if "고렴" in loc:
        return "고렴창고"
    if "판교" in loc:
        return "판교연구소"
    if "석정리" in loc:
        return "석정리창고"
    if "전 공장" in loc or "전공장" in loc:
        return "기타(전공장)"
    if "복지관" in loc:
        return "기타(복지관)"
    return "기타(전공장)"


def extract_team(location_group: str, month: str = "") -> str:
    """소분류 장소 그룹에서 담당 팀을 반환"""
    if location_group in ("평택1공장", "평택2공장"):
        return "환경안전2팀"
    if location_group == "고렴창고":
        # 1월은 환경안전1팀, 나머지는 환경안전2팀
        if month == "1월":
            return "환경안전1팀"
        return "환경안전2팀"
    return "환경안전1팀"


def extract_location_major(location_group: str) -> str:
    """소분류 장소 그룹에서 대분류를 반환"""
    if location_group.startswith("화성"):
        return "화성"
    if location_group.startswith("평택"):
        return "평택"
    if location_group.startswith("고렴"):
        return "고렴"
    if location_group.startswith("판교"):
        return "판교"
    if location_group.startswith("석정리"):
        return "석정리"
    return "기타(전공장)"




def extract_excel_images(file_source: Union[str, io.BytesIO]) -> dict[str, dict[int, str]]:
    """
    ZIP + XML 기반으로 엑셀 내 이미지를 추출.
    Microsoft 365 richData 형식 (셀 내 이미지) 지원.
    file_source: 파일 경로(str) 또는 BytesIO 객체.
    Returns {sheet_name: {row_number(1-based): {"before": url, "after": url}}}.
    """
    result: dict[str, dict[int, dict[str, str]]] = {}
    NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    NS_S = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    NS_RD = "http://schemas.microsoft.com/office/spreadsheetml/2017/richdata"
    NS_RVREL = "http://schemas.microsoft.com/office/spreadsheetml/2022/richvaluerel"

    try:
        with zipfile.ZipFile(file_source, 'r') as zf:
            namelist = set(zf.namelist())

            # 1) Read all media files
            media_data: dict[str, bytes] = {}
            for name in namelist:
                if '/media/' in name:
                    media_data[name.split('/')[-1]] = zf.read(name)
            if not media_data:
                return result

            # 2) workbook.xml.rels → rId to sheet path
            rid_to_path: dict[str, str] = {}
            wb_rels = 'xl/_rels/workbook.xml.rels'
            if wb_rels in namelist:
                for rel in ET.fromstring(zf.read(wb_rels)):
                    rid = rel.get('Id', '')
                    target = rel.get('Target', '')
                    if rid and target:
                        if target.startswith('/'):
                            rid_to_path[rid] = target.lstrip('/')
                        else:
                            rid_to_path[rid] = 'xl/' + target.lstrip('./')

            # 3) workbook.xml → sheet name to sheet path
            sheet_files: dict[str, str] = {}
            if 'xl/workbook.xml' in namelist:
                wb_root = ET.fromstring(zf.read('xl/workbook.xml'))
                for el in wb_root.iter(f'{{{NS_S}}}sheet'):
                    sname = el.get('name', '')
                    rid = el.get(f'{{{NS_R}}}id', '')
                    if sname and rid and rid in rid_to_path:
                        sheet_files[sname] = rid_to_path[rid]

            # 4) Try richData format (Microsoft 365 "Place in Cell" images)
            richdata_rels = 'xl/richData/_rels/richValueRel.xml.rels'
            richdata_rel = 'xl/richData/richValueRel.xml'
            richdata_rv = 'xl/richData/rdrichvalue.xml'

            if all(f in namelist for f in (richdata_rels, richdata_rel, richdata_rv)):
                # 4a) richValueRel.xml.rels → rId to media filename
                rid_to_media: dict[str, str] = {}
                for rel in ET.fromstring(zf.read(richdata_rels)):
                    rid = rel.get('Id', '')
                    target = rel.get('Target', '')
                    if rid and target:
                        rid_to_media[rid] = target.split('/')[-1]

                # 4b) richValueRel.xml → ordered list of rIds
                rvrel_root = ET.fromstring(zf.read(richdata_rel))
                rel_rids: list[str] = []
                for rel_el in rvrel_root:
                    rid = rel_el.get(f'{{{NS_R}}}id', '')
                    rel_rids.append(rid)

                # 4c) rdrichvalue.xml → rv index to media filename
                #     rv[i].v[0] = index into rel_rids
                rv_root = ET.fromstring(zf.read(richdata_rv))
                vm_to_media: dict[int, str] = {}  # vm (1-based) → media filename
                for i, rv in enumerate(rv_root.findall(f'{{{NS_RD}}}rv')):
                    vals = [v.text for v in rv.findall(f'{{{NS_RD}}}v')]
                    if vals:
                        try:
                            rel_idx = int(vals[0])
                            if 0 <= rel_idx < len(rel_rids):
                                rid = rel_rids[rel_idx]
                                media_name = rid_to_media.get(rid, '')
                                if media_name:
                                    vm_to_media[i + 1] = media_name  # vm is 1-based
                        except (ValueError, IndexError):
                            pass

                # 4d) Parse each sheet XML for cells with vm attribute
                # Column N = 개선 전 사진, Column W = 개선 후 사진
                col_to_key = {"N": "before", "W": "after"}
                for sheet_name, sheet_path in sheet_files.items():
                    if sheet_path not in namelist:
                        continue
                    sheet_root = ET.fromstring(zf.read(sheet_path))
                    row_images: dict[int, dict[str, str]] = {}

                    for cell in sheet_root.iter(f'{{{NS_S}}}c'):
                        vm = cell.get('vm')
                        if vm is None:
                            continue
                        ref = cell.get('r', '')
                        col = ''.join(c for c in ref if c.isalpha())
                        img_key = col_to_key.get(col)
                        if not img_key:
                            continue

                        row_num = int(''.join(c for c in ref if c.isdigit()))
                        vm_idx = int(vm)
                        media_name = vm_to_media.get(vm_idx, '')
                        if not media_name or media_name not in media_data:
                            continue
                        if row_num in row_images and img_key in row_images[row_num]:
                            continue

                        # Convert to base64 data URL
                        ext = os.path.splitext(media_name)[1].lower() or '.png'
                        mime = {'.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.gif': 'image/gif', '.webp': 'image/webp'}.get(ext, 'image/png')
                        b64 = base64.b64encode(media_data[media_name]).decode('ascii')
                        data_url = f"data:{mime};base64,{b64}"
                        if row_num not in row_images:
                            row_images[row_num] = {}
                        row_images[row_num][img_key] = data_url

                    if row_images:
                        result[sheet_name] = row_images

    except Exception as e:
        print(f"[extract_excel_images] error: {e}")
        import traceback
        traceback.print_exc()

    return result


def parse_excel(file_source: Union[str, io.BytesIO]) -> list[dict]:
    # ZIP 기반으로 이미지 먼저 추출 (openpyxl 의존 X)
    if isinstance(file_source, io.BytesIO):
        file_source.seek(0)
    all_images = extract_excel_images(file_source)

    if isinstance(file_source, io.BytesIO):
        file_source.seek(0)
    wb = openpyxl.load_workbook(file_source, read_only=True, data_only=True)
    all_records = []

    target_sheets = [s for s in wb.sheetnames if s != "미완료"]

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        month_label = sheet_name
        image_map = all_images.get(sheet_name, {})

        for row_num, row in enumerate(
            ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True),
            start=7,
        ):
            if not row or len(row) < 27:
                continue

            no = row[0]
            if no is None or str(no).strip() == "":
                continue
            try:
                int(no)
            except (ValueError, TypeError):
                continue

            department = str(row[1] or "").strip()
            person = str(row[2] or "").strip()

            if not department and not person:
                continue

            date_val = parse_date(row[3])
            # 월을 date에서 추출 (예: "2025-08-31" → "8월")
            if date_val and len(date_val) >= 7:
                try:
                    month_label = str(int(date_val[5:7])) + "월"
                except ValueError:
                    pass
            location = str(row[4] or "").strip()
            content = str(row[5] or "").strip()
            process = str(row[6] or "").strip()
            disaster_type = str(row[7] or "").strip()

            likelihood_before = parse_number(row[8])
            severity_before = parse_number(row[9])
            risk_before = parse_number(row[10])
            grade_before = str(row[11] or "").strip().replace(" ", "")

            improvement_needed = str(row[12] or "").strip()
            improvement_plan = str(row[14] or "").strip()
            improve_dept = str(row[15] or "").strip()
            planned_date = parse_date(row[16])
            actual_date = parse_date(row[17])

            likelihood_after = parse_number(row[18])
            severity_after = parse_number(row[19])
            risk_after = parse_number(row[20])
            grade_after = str(row[21] or "").strip().replace(" ", "")

            completion = str(row[23] or "").strip()
            note = str(row[24] or "").strip()
            tracking_manager = str(row[25] or "").strip()
            week = parse_number(row[26]) if len(row) > 26 else 0

            if grade_before == "-" or grade_before == "":
                grade_before = "-"

            record = {
                "no": int(no),
                "month": month_label,
                "department": department,
                "person": person,
                "date": date_val,
                "location": location,
                "location_group": extract_location_group(location),
                "location_major": extract_location_major(extract_location_group(location)),
                "content": content[:100],
                "content_full": content,
                "process": process,
                "disaster_type": disaster_type,
                "likelihood_before": likelihood_before,
                "severity_before": severity_before,
                "risk_before": risk_before,
                "grade_before": grade_before,
                "improvement_needed": improvement_needed,
                "improvement_plan": improvement_plan,
                "improve_dept": improve_dept,
                "planned_date": planned_date,
                "actual_date": actual_date,
                "likelihood_after": likelihood_after,
                "severity_after": severity_after,
                "risk_after": risk_after,
                "grade_after": grade_after,
                "completion": completion,
                "note": note,
                "tracking_manager": tracking_manager,
                "week": week,
                "image": (image_map.get(row_num) or {}).get("before", ""),
                "image_after": (image_map.get(row_num) or {}).get("after", ""),
            }
            all_records.append(record)

    wb.close()
    return all_records


CHANNELS = [
    "정기위험성평가(코스맥스)",
    "정기위험성평가(협력사)",
    "수시위험성평가",
    "안전점검",
    "부서별 위험요소발굴",
    "근로자 제안",
    "5S/EHS평가",
]


@app.post("/api/upload")
async def upload_excel(request: Request, file: UploadFile = File(...), channel: str = Form("안전점검")):
    verify_token(request)

    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx, .xlsm)만 업로드 가능합니다.")

    try:
        content = await file.read()
        print(f"[upload] file={file.filename}, size={len(content)} bytes, channel={channel}")
    except Exception as e:
        print(f"[upload] 파일 읽기 오류: {e}")
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"파일 읽기 오류: {str(e)}")

    try:
        file_buf = io.BytesIO(content)
        records = parse_excel(file_buf)
        print(f"[upload] 파싱 완료: {len(records)}건")
    except Exception as e:
        print(f"[upload] 엑셀 파싱 오류: {e}")
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=400, detail=f"엑셀 파싱 오류: {str(e)}")

    # 엑셀 원본 파일 보관 (DB에 바이너리 저장)
    try:
        print(f"[upload] 엑셀 원본 보관 생략 (DB 레코드 기반 다운로드 사용)")
    except Exception as e:
        pass

    for r in records:
        r["channel"] = channel
        r["source"] = "excel"
        r["_id"] = uuid.uuid4().hex
        if not r.get("image"):
            r["image"] = ""
        if not r.get("image_after"):
            r["image_after"] = ""

    try:
        existing = load_data()
        # 엑셀 데이터만 교체, 직접입력(manual) 데이터는 보존
        existing = [r for r in existing if r.get("channel") != channel or r.get("source") == "manual"]
        existing.extend(records)

        save_data(existing)
        print(f"[upload] 저장 완료: 전체 {len(existing)}건")
    except Exception as e:
        print(f"[upload] DB 저장 오류: {e}")
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"데이터 저장 오류: {str(e)}")

    return {"message": f"[{channel}] {len(records)}건 업로드 완료 (전체 {len(existing)}건)", "count": len(records)}


@app.get("/api/channels")
async def get_channels(request: Request):
    verify_token(request)
    return {"channels": CHANNELS}


@app.get("/api/channels/status")
async def channel_status(request: Request):
    verify_token(request)
    data = load_data()
    counts: dict[str, int] = {}
    for r in data:
        ch = r.get("channel", "미분류")
        counts[ch] = counts.get(ch, 0) + 1
    return {"channels": CHANNELS, "counts": counts, "total": len(data)}


@app.post("/api/channels/delete")
async def delete_channel_data(request: Request):
    verify_token(request)
    body = await request.json()
    channel = body.get("channel")
    if not channel:
        raise HTTPException(status_code=400, detail="채널명이 필요합니다.")
    data = load_data()
    before = len(data)
    data = [r for r in data if r.get("channel") != channel]
    after = len(data)
    save_data(data)
    return {"message": f"[{channel}] {before - after}건 삭제 완료", "remaining": after}


@app.post("/api/data/clear")
async def clear_all_data(request: Request):
    verify_token(request)
    save_data([])
    return {"message": "전체 데이터 삭제 완료"}


# --- Image Upload ---
ALLOWED_IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".gif", ".webp", ".heic")

@app.post("/api/image/upload")
async def upload_image(request: Request, file: UploadFile = File(...)):
    verify_token(request)
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        raise HTTPException(status_code=400, detail="이미지 파일만 업로드 가능합니다. (jpg, png, gif, webp)")
    content = await file.read()
    mime = {'.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.gif': 'image/gif', '.webp': 'image/webp', '.heic': 'image/heic'}.get(ext, 'image/png')
    b64 = base64.b64encode(content).decode('ascii')
    data_url = f"data:{mime};base64,{b64}"
    return {"filename": file.filename, "url": data_url}


@app.get("/uploads/images/{filename}")
async def get_image(filename: str):
    filepath = os.path.join(IMAGE_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="이미지를 찾을 수 없습니다.")
    return FileResponse(filepath)


# --- Direct Record Input ---
@app.post("/api/record/add")
async def add_record(request: Request):
    verify_token(request)
    body = await request.json()

    # Required fields
    channel = body.get("channel", "").strip()
    month = body.get("month", "").strip()
    person = body.get("person", "").strip()
    date_val = body.get("date", "").strip()
    location = body.get("location", "").strip()
    workplace = body.get("workplace", "").strip()
    content = body.get("content", "").strip()
    cause_object = body.get("cause_object", "").strip()
    process = body.get("process", "").strip()
    disaster_type = body.get("disaster_type", "").strip()
    improvement_plan = body.get("improvement_plan", "").strip()
    completion = body.get("completion", "미완료").strip()
    week = parse_number(body.get("week", 0))
    image = body.get("image", "").strip()
    image_after = body.get("image_after", "").strip()

    likelihood_before = parse_number(body.get("likelihood_before", 0))
    severity_before = parse_number(body.get("severity_before", 0))
    risk_before = likelihood_before * severity_before
    grade_before = "A" if risk_before <= 4 else "B" if risk_before <= 8 else "C" if risk_before <= 12 else "D" if risk_before > 0 else "-"

    likelihood_after = parse_number(body.get("likelihood_after", 0))
    severity_after = parse_number(body.get("severity_after", 0))
    risk_after = likelihood_after * severity_after
    grade_after = "A" if risk_after <= 4 else "B" if risk_after <= 8 else "C" if risk_after <= 12 else "D" if risk_after > 0 else "-"

    if not channel or not content:
        raise HTTPException(status_code=400, detail="구분(채널)과 위험요소 내용은 필수입니다.")

    existing = load_data()
    max_no = max((r.get("no", 0) for r in existing), default=0)

    record = {
        "_id": uuid.uuid4().hex,
        "no": max_no + 1,
        "month": month,
        "department": "",
        "person": person,
        "date": parse_date(date_val),
        "location": location,
        "workplace": workplace,
        "location_group": extract_location_group(location),
        "location_major": extract_location_major(extract_location_group(location)),
        "content": content[:100],
        "content_full": content,
        "cause_object": cause_object,
        "process": process,
        "disaster_type": disaster_type,
        "likelihood_before": likelihood_before,
        "severity_before": severity_before,
        "risk_before": risk_before,
        "grade_before": grade_before,
        "improvement_needed": "",
        "improvement_plan": improvement_plan,
        "improve_dept": "",
        "planned_date": None,
        "actual_date": None,
        "likelihood_after": likelihood_after,
        "severity_after": severity_after,
        "risk_after": risk_after,
        "grade_after": grade_after,
        "completion": completion,
        "note": "",
        "tracking_manager": "",
        "week": week,
        "channel": channel,
        "source": "manual",
        "image": image,
        "image_after": image_after,
    }

    existing.append(record)
    save_data(existing)

    return {"message": f"위험요소 1건 추가 완료 (No.{record['no']})", "record": record}


@app.post("/api/record/update")
async def update_record(request: Request):
    verify_token(request)
    body = await request.json()
    record_id = body.get("_id", "").strip()
    if not record_id:
        raise HTTPException(status_code=400, detail="_id가 필요합니다.")

    data = load_data()
    target = None
    for r in data:
        if r.get("_id") == record_id:
            target = r
            break
    if not target:
        raise HTTPException(status_code=404, detail="레코드를 찾을 수 없습니다.")

    # Update editable fields
    for field in ("channel", "month", "person", "date", "location", "workplace", "content",
                  "cause_object", "process", "disaster_type", "improvement_plan", "completion", "image", "image_after"):
        if field in body:
            target[field] = body[field].strip() if isinstance(body[field], str) else body[field]

    if "date" in body:
        target["date"] = parse_date(body["date"])
    if "location" in body:
        target["location"] = body["location"].strip()
        target["location_group"] = extract_location_group(target["location"])
        target["location_major"] = extract_location_major(target["location_group"])
    if "content" in body:
        target["content_full"] = body["content"].strip()
        target["content"] = body["content"].strip()[:100]
    if "week" in body:
        target["week"] = parse_number(body["week"])

    if "likelihood_before" in body or "severity_before" in body:
        lh = parse_number(body.get("likelihood_before", target.get("likelihood_before", 0)))
        sv = parse_number(body.get("severity_before", target.get("severity_before", 0)))
        target["likelihood_before"] = lh
        target["severity_before"] = sv
        target["risk_before"] = lh * sv
        risk = lh * sv
        target["grade_before"] = "A" if risk <= 4 else "B" if risk <= 8 else "C" if risk <= 12 else "D" if risk > 0 else "-"

    if "likelihood_after" in body or "severity_after" in body:
        lh = parse_number(body.get("likelihood_after", target.get("likelihood_after", 0)))
        sv = parse_number(body.get("severity_after", target.get("severity_after", 0)))
        target["likelihood_after"] = lh
        target["severity_after"] = sv
        target["risk_after"] = lh * sv
        risk = lh * sv
        target["grade_after"] = "A" if risk <= 4 else "B" if risk <= 8 else "C" if risk <= 12 else "D" if risk > 0 else "-"

    save_data(data)

    return {"message": "수정 완료", "record": target}


@app.post("/api/record/delete")
async def delete_record(request: Request):
    verify_token(request)
    body = await request.json()
    record_id = body.get("_id", "").strip()
    if not record_id:
        raise HTTPException(status_code=400, detail="_id가 필요합니다.")

    data = load_data()
    # 삭제 전 레코드 보관 (엑셀 동기화용)
    deleted_record = None
    for r in data:
        if r.get("_id") == record_id:
            deleted_record = r
            break
    if not deleted_record:
        raise HTTPException(status_code=404, detail="레코드를 찾을 수 없습니다.")

    data = [r for r in data if r.get("_id") != record_id]
    save_data(data)

    return {"message": "삭제 완료"}


# --- Data API ---
def _cleanup_records(data: list[dict]) -> tuple[list[dict], list[str]]:
    """Clean up records and return (data, dirty_ids)."""
    dirty_ids = []
    for r in data:
        changed = False
        if "channel" not in r:
            r["channel"] = "안전점검"
            changed = True
        if "_id" not in r:
            r["_id"] = uuid.uuid4().hex
            changed = True
        if "image" not in r:
            r["image"] = ""
            changed = True
        if "image_after" not in r:
            r["image_after"] = ""
            changed = True
        new_lg = extract_location_group(r.get("location", ""))
        if r.get("location_group") != new_lg:
            r["location_group"] = new_lg
            changed = True
        new_lm = extract_location_major(new_lg)
        if r.get("location_major") != new_lm:
            r["location_major"] = new_lm
            changed = True
        if changed:
            dirty_ids.append(r["_id"])
    return data, dirty_ids


def invalidate_data_cache():
    global _data_cache, _data_cache_time
    _data_cache = None
    _data_cache_time = 0

def load_data() -> list[dict]:
    global _data_cache, _data_cache_time
    now = _time.time()
    if _data_cache is not None and (now - _data_cache_time) < _DATA_CACHE_TTL:
        return [dict(r) for r in _data_cache]  # shallow copy

    if not DATABASE_URL:
        if not os.path.exists(DATA_FILE):
            return []
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        data, dirty_ids = _cleanup_records(data)
        if dirty_ids:
            with open(DATA_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        _data_cache = data
        _data_cache_time = now
        return [dict(r) for r in data]

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT data FROM records")
    rows = cur.fetchall()
    data = [row[0] for row in rows]
    data, dirty_ids = _cleanup_records(data)
    if dirty_ids:
        dirty_set = set(dirty_ids)
        for r in data:
            if r["_id"] in dirty_set:
                cur.execute("UPDATE records SET data = %s WHERE _id = %s",
                            (json.dumps(r, ensure_ascii=False), r["_id"]))
        conn.commit()
    cur.close()
    release_db(conn)
    _data_cache = data
    _data_cache_time = now
    return data


def save_data(data: list[dict]):
    invalidate_data_cache()
    if not DATABASE_URL:
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return

    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM records")
    if data:
        args = []
        for r in data:
            rid = r.get("_id") or uuid.uuid4().hex
            r["_id"] = rid
            args.append((rid, json.dumps(r, ensure_ascii=False)))
        psycopg2.extras.execute_values(
            cur, "INSERT INTO records (_id, data) VALUES %s", args, template="(%s, %s)", page_size=500
        )
    conn.commit()
    cur.close()
    release_db(conn)


# --- Excel File Storage ---
def save_excel_file(channel: str, file_path: str, filename: str):
    """엑셀 파일을 DB 또는 파일시스템에 보관"""
    with open(file_path, "rb") as f:
        data = f.read()
    if DATABASE_URL:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO excel_files (channel, filename, data, updated_at) VALUES (%s, %s, %s, %s) "
            "ON CONFLICT (channel) DO UPDATE SET filename = %s, data = %s, updated_at = %s",
            (channel, filename, data, datetime.now().isoformat(),
             filename, data, datetime.now().isoformat())
        )
        conn.commit()
        cur.close()
        release_db(conn)
    else:
        safe_name = channel.replace("/", "_").replace(" ", "_")
        dest = os.path.join(UPLOAD_DIR, f"excel_{safe_name}.xlsm")
        with open(dest, "wb") as f:
            f.write(data)


def load_excel_file(channel: str) -> Optional[bytes]:
    """보관된 엑셀 파일 바이너리 반환"""
    if DATABASE_URL:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT data, filename FROM excel_files WHERE channel = %s", (channel,))
        row = cur.fetchone()
        cur.close()
        release_db(conn)
        if row:
            return row[0]
        return None
    else:
        safe_name = channel.replace("/", "_").replace(" ", "_")
        path = os.path.join(UPLOAD_DIR, f"excel_{safe_name}.xlsm")
        if os.path.exists(path):
            with open(path, "rb") as f:
                return f.read()
        return None


def load_excel_filename(channel: str) -> Optional[str]:
    """보관된 엑셀 파일명 반환"""
    if DATABASE_URL:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT filename FROM excel_files WHERE channel = %s", (channel,))
        row = cur.fetchone()
        cur.close()
        release_db(conn)
        return row[0] if row else None
    return None


def save_excel_bytes(channel: str, data: bytes, filename: str = "updated.xlsm"):
    """수정된 엑셀 바이너리를 저장"""
    if DATABASE_URL:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO excel_files (channel, filename, data, updated_at) VALUES (%s, %s, %s, %s) "
            "ON CONFLICT (channel) DO UPDATE SET data = %s, updated_at = %s",
            (channel, filename, data, datetime.now().isoformat(),
             data, datetime.now().isoformat())
        )
        conn.commit()
        cur.close()
        release_db(conn)
    else:
        safe_name = channel.replace("/", "_").replace(" ", "_")
        dest = os.path.join(UPLOAD_DIR, f"excel_{safe_name}.xlsm")
        with open(dest, "wb") as f:
            f.write(data)


# --- Excel Write-Back ---
# 엑셀 컬럼 매핑 (0-based index → record field)
EXCEL_COL_MAP = {
    0: "no", 1: "department", 2: "person", 3: "date", 4: "location",
    5: "content_full", 6: "process", 7: "disaster_type",
    8: "likelihood_before", 9: "severity_before", 10: "risk_before", 11: "grade_before",
    12: "improvement_needed", 14: "improvement_plan", 15: "improve_dept",
    16: "planned_date", 17: "actual_date",
    18: "likelihood_after", 19: "severity_after", 20: "risk_after", 21: "grade_after",
    23: "completion", 24: "note", 25: "tracking_manager", 26: "week",
}

# 역매핑: field → col index
FIELD_TO_COL = {v: k for k, v in EXCEL_COL_MAP.items()}


def _update_excel_row(ws, row_num: int, record: dict):
    """엑셀 워크시트의 특정 행을 record 데이터로 업데이트"""
    for col_idx, field in EXCEL_COL_MAP.items():
        val = record.get(field, "")
        if val is None:
            val = ""
        ws.cell(row=row_num, column=col_idx + 1, value=val)


def _find_excel_row_by_no(ws, no_val: int, min_row: int = 7) -> Optional[int]:
    """엑셀에서 No 값으로 행 번호를 찾음"""
    for row_num in range(min_row, ws.max_row + 1):
        cell_val = ws.cell(row=row_num, column=1).value
        if cell_val is not None:
            try:
                if int(cell_val) == no_val:
                    return row_num
            except (ValueError, TypeError):
                continue
    return None


def _get_next_empty_row(ws, min_row: int = 7) -> int:
    """데이터가 있는 마지막 행 다음 행 번호 반환"""
    last_row = min_row
    for row_num in range(min_row, ws.max_row + 1):
        cell_val = ws.cell(row=row_num, column=1).value
        if cell_val is not None:
            try:
                int(cell_val)
                last_row = row_num + 1
            except (ValueError, TypeError):
                continue
    return last_row


def sync_record_to_excel(channel: str, record: dict, action: str = "upsert"):
    """
    보관된 엑셀 파일에 레코드를 동기화.
    action: "upsert" (추가/수정), "delete" (삭제)
    """
    import io
    excel_data = load_excel_file(channel)
    if not excel_data:
        return  # 보관된 엑셀 없으면 skip

    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_data))
    except Exception as e:
        print(f"[excel-sync] 엑셀 열기 오류: {e}")
        return

    # 월 기반으로 시트 찾기, 없으면 첫 번째 시트
    month = record.get("month", "")
    target_ws = None
    for sname in wb.sheetnames:
        if sname == month or month in sname:
            target_ws = wb[sname]
            break
    if not target_ws:
        # 미완료 시트 제외, 첫 번째 시트 사용
        for sname in wb.sheetnames:
            if sname != "미완료":
                target_ws = wb[sname]
                break
    if not target_ws:
        wb.close()
        return

    no_val = record.get("no")
    if no_val is None:
        wb.close()
        return

    if action == "delete":
        row_num = _find_excel_row_by_no(target_ws, int(no_val))
        if row_num:
            # 행의 모든 셀을 비움
            for col in range(1, 28):
                target_ws.cell(row=row_num, column=col, value=None)
    else:  # upsert
        row_num = _find_excel_row_by_no(target_ws, int(no_val))
        if row_num:
            _update_excel_row(target_ws, row_num, record)
        else:
            new_row = _get_next_empty_row(target_ws)
            _update_excel_row(target_ws, new_row, record)

    # 저장
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    filename = load_excel_filename(channel) or "updated.xlsm"
    save_excel_bytes(channel, buf.getvalue(), filename)


@app.get("/api/data")
async def get_data(request: Request):
    verify_token(request)
    records = load_data()
    return {"records": records, "total": len(records)}


@app.get("/api/record-image/{record_id}")
async def get_record_image(request: Request, record_id: str, field: str = "image"):
    verify_token(request)
    if DATABASE_URL:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT data FROM records WHERE _id = %s", (record_id,))
        row = cur.fetchone()
        cur.close()
        release_db(conn)
        if row:
            r = json.loads(row[0]) if isinstance(row[0], str) else row[0]
            img = r.get(field, "")
            if img:
                return {"url": img}
    return {"url": ""}


@app.get("/api/summary")
async def get_summary(
    request: Request,
    channel: Optional[str] = None,
    year: Optional[str] = None,
    month: Optional[str] = None,
    location: Optional[str] = None,
    grade: Optional[str] = None,
    disaster_type: Optional[str] = None,
    process: Optional[str] = None,
    person: Optional[str] = None,
    week: Optional[int] = None,
    keyword: Optional[str] = None,
    completion: Optional[str] = None,
    team: Optional[str] = None,
    page: Optional[str] = None,
    sel_week: Optional[int] = None,
):
    verify_token(request)
    records = load_data()
    _all_records_cache = list(records)  # 필터 옵션용 원본 보관 (load_data 이중 호출 방지)

    # Apply filters
    if team and team != "전체":
        records = [r for r in records if extract_team(r.get("location_group", ""), r.get("month", "")) == team]
    if channel and channel != "전체":
        records = [r for r in records if r.get("channel") == channel]
    if year and year != "전체":
        def match_year(r):
            d = r.get("date") or ""
            if len(d) >= 4 and d[:4] == year:
                return True
            # date가 없는 레코드는 upload_year 또는 제외하지 않음
            if not d:
                return True
            return False
        records = [r for r in records if match_year(r)]
    if month and month != "전체":
        records = [r for r in records if r["month"] == month]
    if location and location != "전체":
        records = [r for r in records if r["location_group"] == location]
    if grade and grade != "전체":
        records = [r for r in records if r["grade_before"] == grade]
    if disaster_type and disaster_type != "전체":
        records = [r for r in records if r["disaster_type"] == disaster_type]
    if process and process != "전체":
        records = [r for r in records if r["process"] == process]
    if person and person != "전체":
        records = [r for r in records if r["person"] == person]
    if week and week > 0:
        records = [r for r in records if r["week"] == week]
    if completion and completion != "전체":
        records = [r for r in records if r["completion"] == completion]
    if keyword:
        kw = keyword.lower()
        records = [r for r in records if kw in r.get("content_full", "").lower()
                   or kw in r.get("location", "").lower()
                   or kw in r.get("improvement_plan", "").lower()]

    # Detect repeated risks by normalized content

    def normalize_content(text):
        if not text:
            return ""
        t = text.strip()
        t = re.sub(r'[.,!?;:\-~·…\s]+', '', t)
        return t.lower()

    content_counts = Counter()
    norm_map = {}
    for r in records:
        c = r.get("content_full", "").strip()
        norm = normalize_content(c)
        if norm:
            content_counts[norm] += 1
            norm_map[id(r)] = norm

    # Mark repeat info on each record
    for r in records:
        norm = norm_map.get(id(r), "")
        cnt = content_counts.get(norm, 0)
        r["repeat_count"] = cnt
        r["is_repeat"] = cnt >= 2

    total = len(records)
    repeat_total = sum(1 for r in records if r["is_repeat"])

    # Grade counts (before improvement)
    grade_a = sum(1 for r in records if r["grade_before"] == "A")
    grade_b = sum(1 for r in records if r["grade_before"] == "B")
    grade_c = sum(1 for r in records if r["grade_before"] == "C")
    grade_d = sum(1 for r in records if r["grade_before"] == "D")

    grade_a_after = sum(1 for r in records if r.get("grade_after") == "A")
    grade_b_after = sum(1 for r in records if r.get("grade_after") == "B")
    grade_c_after = sum(1 for r in records if r.get("grade_after") == "C")
    grade_d_after = sum(1 for r in records if r.get("grade_after") == "D")

    # 현황 등급: 완료 건은 grade_after, 미완료 건은 grade_before
    def current_grade(r):
        if r["completion"] == "완료" and r.get("grade_after") in ("A", "B", "C", "D"):
            return r["grade_after"]
        return r.get("grade_before", "-")
    grade_a_current = sum(1 for r in records if current_grade(r) == "A")
    grade_b_current = sum(1 for r in records if current_grade(r) == "B")
    grade_c_current = sum(1 for r in records if current_grade(r) == "C")
    grade_d_current = sum(1 for r in records if current_grade(r) == "D")

    complete = sum(1 for r in records if r["completion"] == "완료")
    incomplete = sum(1 for r in records if r["completion"] != "완료")
    improvement_rate = round(complete / total * 100, 1) if total > 0 else 0

    # Shared month sort helper
    def month_sort_key(m):
        try:
            return int(m.replace("월", ""))
        except (ValueError, AttributeError):
            return 0
    all_months = sorted(set(r["month"] for r in records), key=month_sort_key)

    need_analysis = page in (None, "summary", "analysis")
    # --- Heavy analysis stats (skip for records page) ---
    if need_analysis:
        grade_cumulative = {}
        cumul = {"A": 0, "B": 0, "C": 0, "D": 0}
        cumul_total = 0
        cumul_complete = 0
        for m in all_months:
            month_recs = [r for r in records if r["month"] == m]
            for r in month_recs:
                g = r["grade_before"] if r["grade_before"] in ("A", "B", "C", "D") else None
                if g and r["completion"] != "완료":
                    cumul[g] += 1
            cumul_total += len(month_recs)
            cumul_complete += sum(1 for r in month_recs if r["completion"] == "완료")
            grade_cumulative[m] = {
                "A": cumul["A"], "B": cumul["B"], "C": cumul["C"], "D": cumul["D"],
                "total_remaining": cumul_total - cumul_complete,
            }

        d_grade_total = 0
        d_after = {"A": 0, "B": 0, "C": 0, "D": 0, "미완료": 0}
        for r in records:
            if r["grade_before"] == "D":
                d_grade_total += 1
                ga = r.get("grade_after")
                if ga in ("A", "B", "C", "D"):
                    d_after[ga] += 1
                else:
                    d_after["미완료"] += 1

        d_grade_monthly = {}
        for m in all_months:
            month_d_recs = [r for r in records if r["month"] == m and r["grade_before"] == "D"]
            before_count = len(month_d_recs)
            after = {"A": 0, "B": 0, "C": 0, "D": 0, "미완료": 0}
            for r in month_d_recs:
                ga = r.get("grade_after")
                if ga in ("A", "B", "C", "D"):
                    after[ga] += 1
                else:
                    after["미완료"] += 1
            d_grade_monthly[m] = {"before": before_count, "after": after}

        def grade_to_num(g):
            return {"A": 1, "B": 2, "C": 3, "D": 4}.get(g, 0)

        risk_trend: dict[str, dict[str, float]] = {}
        for m in all_months:
            month_recs = [r for r in records if r["month"] == m]
            before_scores = [r["risk_before"] for r in month_recs if r["risk_before"] and r["risk_before"] > 0]
            after_scores = [r["risk_after"] for r in month_recs if r["risk_after"] and r["risk_after"] > 0]
            before_grades = [grade_to_num(r["grade_before"]) for r in month_recs if r["grade_before"] in ("A","B","C","D")]
            after_grades = [grade_to_num(r.get("grade_after","")) for r in month_recs if r.get("grade_after") in ("A","B","C","D")]
            risk_trend[m] = {
                "avg_before": round(sum(before_scores) / len(before_scores), 1) if before_scores else 0,
                "avg_after": round(sum(after_scores) / len(after_scores), 1) if after_scores else 0,
                "avg_grade_before": round(sum(before_grades) / len(before_grades), 2) if before_grades else 0,
                "avg_grade_after": round(sum(after_grades) / len(after_grades), 2) if after_grades else 0,
            }

        monthly_effort = {}
        for m in all_months:
            month_recs = [r for r in records if r["month"] == m]
            found = len(month_recs)
            completed = sum(1 for r in month_recs if r["completion"] == "완료")
            rate = round(completed / found * 100, 1) if found > 0 else 0
            monthly_effort[m] = {"found": found, "completed": completed, "rate": rate}

        location_stats: dict[str, dict[str, int]] = {}
        for r in records:
            lg = r["location_group"]
            if lg not in location_stats:
                location_stats[lg] = {"A": 0, "B": 0, "C": 0, "D": 0, "-": 0}
            g = r["grade_before"] if r["grade_before"] in ("A", "B", "C", "D") else "-"
            location_stats[lg][g] += 1

        location_disaster_stats: dict[str, dict[str, int]] = {}
        all_disaster_types_set: set[str] = set()
        for r in records:
            lg = r["location_group"]
            dt = r["disaster_type"] if r["disaster_type"] else "미분류"
            all_disaster_types_set.add(dt)
            if lg not in location_disaster_stats:
                location_disaster_stats[lg] = {}
            location_disaster_stats[lg][dt] = location_disaster_stats[lg].get(dt, 0) + 1

        MAJOR_ORDER = ["화성", "평택", "고렴", "판교"]
        location_major_stats: dict[str, dict[str, int]] = {}
        location_major_disaster_stats: dict[str, dict[str, int]] = {}
        location_hierarchy: dict[str, list[str]] = {m: [] for m in MAJOR_ORDER}
        for r in records:
            lg = r["location_group"]
            lm = r.get("location_major") or extract_location_major(lg)
            if lm not in location_major_stats:
                location_major_stats[lm] = {"A": 0, "B": 0, "C": 0, "D": 0, "-": 0}
            g = r["grade_before"] if r["grade_before"] in ("A", "B", "C", "D") else "-"
            location_major_stats[lm][g] += 1
            dt = r["disaster_type"] if r["disaster_type"] else "미분류"
            if lm not in location_major_disaster_stats:
                location_major_disaster_stats[lm] = {}
            location_major_disaster_stats[lm][dt] = location_major_disaster_stats[lm].get(dt, 0) + 1
            if lm in location_hierarchy and lg not in location_hierarchy[lm]:
                location_hierarchy[lm].append(lg)
        for m in location_hierarchy:
            location_hierarchy[m].sort()

        grade_trend: dict[str, dict[str, int]] = {}
        grade_trend_after: dict[str, dict[str, int]] = {}
        for r in records:
            m = r["month"]
            if m not in grade_trend:
                grade_trend[m] = {"A": 0, "B": 0, "C": 0, "D": 0}
            g = r["grade_before"] if r["grade_before"] in ("A", "B", "C", "D") else None
            if g:
                grade_trend[m][g] += 1
            if m not in grade_trend_after:
                grade_trend_after[m] = {"A": 0, "B": 0, "C": 0, "D": 0}
            ga = r["grade_after"] if r.get("grade_after") in ("A", "B", "C", "D") else None
            if ga:
                grade_trend_after[m][ga] += 1
        grade_trend = dict(sorted(grade_trend.items(), key=lambda x: month_sort_key(x[0])))

        week_stats: dict[str, int] = {}
        for r in records:
            m = r["month"]
            w = r["week"]
            if w > 0:
                key = f"{m} {w}주차"
                week_stats[key] = week_stats.get(key, 0) + 1
        def week_sort_key(k):
            parts = k.split()
            try: mon = int(parts[0].replace("월", ""))
            except (ValueError, IndexError): mon = 999
            try: wk = int(parts[1].replace("주차", ""))
            except (ValueError, IndexError): wk = 999
            return (mon, wk)
        week_stats = dict(sorted(week_stats.items(), key=lambda x: week_sort_key(x[0])))

        disaster_stats: dict[str, int] = {}
        for r in records:
            dt = r["disaster_type"] if r["disaster_type"] else "미분류"
            disaster_stats[dt] = disaster_stats.get(dt, 0) + 1

        process_stats: dict[str, int] = {}
        for r in records:
            p = r["process"] if r["process"] else "미분류"
            process_stats[p] = process_stats.get(p, 0) + 1

        channel_stats: dict[str, int] = {}
        for r in records:
            ch = r.get("channel", "미분류")
            channel_stats[ch] = channel_stats.get(ch, 0) + 1

        channel_grade_stats: dict[str, dict[str, int]] = {}
        for r in records:
            ch = r.get("channel", "미분류")
            if ch not in channel_grade_stats:
                channel_grade_stats[ch] = {"A": 0, "B": 0, "C": 0, "D": 0, "-": 0, "complete": 0, "incomplete": 0}
            g = r["grade_before"] if r["grade_before"] in ("A", "B", "C", "D") else "-"
            channel_grade_stats[ch][g] += 1
            if r["completion"] == "완료":
                channel_grade_stats[ch]["complete"] += 1
            else:
                channel_grade_stats[ch]["incomplete"] += 1
    else:
        # records page — skip heavy stats
        grade_cumulative = {}
        d_grade_total = 0
        d_after = {}
        d_grade_monthly = {}
        risk_trend = {}
        monthly_effort = {}
        location_stats = {}
        location_disaster_stats = {}
        location_major_stats = {}
        location_major_disaster_stats = {}
        location_hierarchy = {}
        grade_trend = {}
        grade_trend_after = {}
        week_stats = {}
        disaster_stats = {}
        process_stats = {}
        channel_stats = {}
        channel_grade_stats = {}

    # Filter options (use already loaded records from top — no second load_data call)
    all_records_for_filters = _all_records_cache
    channels = sorted(set(r.get("channel", "미분류") for r in all_records_for_filters))
    years = sorted(set(r["date"][:4] for r in all_records_for_filters if r.get("date") and len(r["date"]) >= 4))
    months = sorted(set(r["month"] for r in all_records_for_filters))
    locations = sorted(set(r["location_group"] for r in all_records_for_filters))
    disaster_types = sorted(set(r["disaster_type"] for r in all_records_for_filters if r["disaster_type"]))
    processes = sorted(set(r["process"] for r in all_records_for_filters if r["process"]))
    persons = sorted(set(r["person"] for r in all_records_for_filters if r["person"]))
    weeks = sorted(set(r["week"] for r in all_records_for_filters if r["week"] > 0))

    # Simple view summary (current filtered records)
    today = date.today()
    cur_month_str = f"{today.month}월"
    cur_week = calc_week_from_date(today.isoformat())
    # 선택된 주차가 있으면 그 주차, 없으면 이번주
    target_week = sel_week if sel_week and sel_week > 0 else cur_week

    view_summary = {
        "total": total, "complete": complete, "incomplete": incomplete,
        "team1": 0, "team1_complete": 0,
        "team2": 0, "team2_complete": 0,
        "week_discovered": 0, "week_improved": 0,
        "week_team1_discovered": 0, "week_team1_improved": 0,
        "week_team2_discovered": 0, "week_team2_improved": 0,
    }
    for r in records:
        tm = extract_team(r.get("location_group", ""), r.get("month", ""))
        is_team1 = (tm == "환경안전1팀")
        if is_team1:
            view_summary["team1"] += 1
            if r["completion"] == "완료":
                view_summary["team1_complete"] += 1
        else:
            view_summary["team2"] += 1
            if r["completion"] == "완료":
                view_summary["team2_complete"] += 1
        # 이번주 판별
        r_date = r.get("date", "")
        r_month = r.get("month", "")
        if r_date and r_date[:4] == str(today.year) and r_month == cur_month_str:
            r_week = r.get("week", 0) or calc_week_from_date(r_date)
            if r_week == target_week:
                view_summary["week_discovered"] += 1
                if is_team1:
                    view_summary["week_team1_discovered"] += 1
                else:
                    view_summary["week_team2_discovered"] += 1
                if r["completion"] == "완료":
                    view_summary["week_improved"] += 1
                    if is_team1:
                        view_summary["week_team1_improved"] += 1
                    else:
                        view_summary["week_team2_improved"] += 1

    # Strip large image data from summary response to reduce payload size
    for r in records:
        r["has_image"] = bool(r.get("image"))
        r["has_image_after"] = bool(r.get("image_after"))
        r.pop("image", None)
        r.pop("image_after", None)

    return {
        "total": total,
        "improvement_rate": improvement_rate,
        "repeat_total": repeat_total,
        "grade_a": grade_a,
        "grade_b": grade_b,
        "grade_c": grade_c,
        "grade_d": grade_d,
        "grade_a_after": grade_a_after,
        "grade_b_after": grade_b_after,
        "grade_c_after": grade_c_after,
        "grade_d_after": grade_d_after,
        "grade_a_current": grade_a_current,
        "grade_b_current": grade_b_current,
        "grade_c_current": grade_c_current,
        "grade_d_current": grade_d_current,
        "grade_cumulative": grade_cumulative,
        "risk_trend": risk_trend,
        "monthly_effort": monthly_effort,
        "d_grade_total": d_grade_total,
        "d_after": d_after,
        "d_grade_monthly": d_grade_monthly,
        "complete": complete,
        "incomplete": incomplete,
        "view_summary": view_summary,
        "location_stats": location_stats,
        "location_disaster_stats": location_disaster_stats,
        "location_major_stats": location_major_stats,
        "location_major_disaster_stats": location_major_disaster_stats,
        "location_hierarchy": location_hierarchy,
        "grade_trend": grade_trend,
        "grade_trend_after": grade_trend_after,
        "week_stats": week_stats,
        "disaster_stats": disaster_stats,
        "process_stats": process_stats,
        "channel_stats": channel_stats,
        "channel_grade_stats": channel_grade_stats,
        "records": records,
        "filters": {
            "channels": channels,
            "years": years,
            "months": months,
            "locations": locations,
            "disaster_types": disaster_types,
            "processes": processes,
            "persons": persons,
            "weeks": weeks,
        },
    }


# --- Weekly Snapshot ---
CHANNEL_GROUPS = {
    "안전점검": ["안전점검"],
    "부서별 위험요소발굴": ["부서별 위험요소발굴"],
    "근로자 제안": ["근로자 제안"],
    "5S/EHS평가": ["5S/EHS평가"],
    "정기위험성평가": ["정기위험성평가(코스맥스)", "정기위험성평가(협력사)"],
    "수시위험성평가": ["수시위험성평가"],
}
CHANNEL_ORDER = ["안전점검", "부서별 위험요소발굴", "근로자 제안", "5S/EHS평가", "정기위험성평가", "수시위험성평가"]
SITE_GROUPS = {
    "전체": None,
    "환경안전1팀": ["화성", "판교"],
    "환경안전2팀": ["평택", "고렴"],
}
QUARTER_MONTHS = {1: [1, 2, 3], 2: [4, 5, 6], 3: [7, 8, 9], 4: [10, 11, 12]}


def calc_week_from_date(date_str: str) -> int:
    """날짜 문자열에서 해당 월의 주차를 계산 (목요일 기준, 1월은 예외)"""
    try:
        from datetime import timedelta
        d = datetime.strptime(date_str, "%Y-%m-%d")
        if d.month == 1:
            # 1월 예외: 1~11일=1주차, 12~18=2주차, 19~25=3주차, 26~=4주차
            if d.day <= 11:
                return 1
            return min((d.day - 12) // 7 + 2, 5)
        # 나머지 월: 해당 날짜가 속한 주의 목요일을 구해서 주차 판단
        diff_to_thu = (3 - d.weekday()) % 7
        if d.weekday() > 3:  # 금,토,일 → 이번 주 목요일은 과거
            diff_to_thu = diff_to_thu - 7
        thu = d + timedelta(days=diff_to_thu)
        if thu.month != d.month:
            if thu.month > d.month or (thu.month == 1 and d.month == 12):
                return 5
            else:
                return 1
        week = (thu.day - 1) // 7 + 1
        return min(week, 5)
    except (ValueError, TypeError):
        return 0


def get_effective_week(r: dict) -> int:
    """레코드의 주차를 반환. week=0이면 날짜 기반으로 자동 계산"""
    w = r.get("week", 0)
    if w and w > 0:
        return w
    date_str = r.get("date", "")
    if date_str:
        return calc_week_from_date(date_str)
    return 0


def compute_quarter_stats(records: list, year: str, quarter: int) -> dict:
    """분기 전체 채널별/사업장별/주차별 발굴·개선 집계 (발굴 주차 기준)"""
    months = QUARTER_MONTHS[quarter]

    # Filter records for this year & quarter months
    filtered = []
    for r in records:
        d = r.get("date", "") or ""
        if d[:4] != year:
            continue
        m_str = r.get("month", "")
        try:
            m_num = int(m_str.replace("월", ""))
        except (ValueError, AttributeError):
            continue
        if m_num in months:
            filtered.append(r)

    # Compute previous year totals
    prev_year = str(int(year) - 1)
    prev_records = [r for r in records if (r.get("date", "") or "")[:4] == prev_year]

    def match_site(r, majors):
        """1월 고렴창고/기타는 화성/판교(환경안전1팀) 소속"""
        lm = r.get("location_major", "")
        is_jan = r.get("month") == "1월"
        # 1월 고렴, 기타(전공장)은 화성/판교 그룹
        if is_jan and lm in ("고렴", "기타(전공장)"):
            return "화성" in majors or "판교" in majors
        if lm in majors:
            return True
        return False

    result = {}
    for site_name, majors in SITE_GROUPS.items():
        site_recs = filtered
        site_prev = prev_records
        if majors:
            site_recs = [r for r in filtered if match_site(r, majors)]
            site_prev = [r for r in prev_records if match_site(r, majors)]

        site_data = {}
        for ch_name in CHANNEL_ORDER:
            ch_keys = CHANNEL_GROUPS[ch_name]
            ch_recs = [r for r in site_recs if r.get("channel", "") in ch_keys]
            ch_prev = [r for r in site_prev if r.get("channel", "") in ch_keys]

            # Previous year totals
            prev_discovered = len(ch_prev)
            prev_improved = sum(1 for r in ch_prev if r.get("completion") == "완료")
            prev_rate = round(prev_improved / prev_discovered, 4) if prev_discovered > 0 else 0

            # Per month/week breakdown
            weeks_data = {}
            for m in months:
                m_key = f"{m}월"
                for w in range(1, 6):
                    w_recs = [r for r in ch_recs if r.get("month") == m_key and get_effective_week(r) == w]
                    if not w_recs and w == 5:
                        continue  # skip 5th week if no data
                    key = f"{m}-{w}"
                    weeks_data[key] = {
                        "discovered": len(w_recs),
                        "d_discovered": sum(1 for r in w_recs if r.get("grade_before") == "D"),
                        "improved": sum(1 for r in w_recs if r.get("completion") == "완료"),
                        "d_improved": sum(1 for r in w_recs if r.get("grade_before") == "D" and r.get("completion") == "완료"),
                    }

            # Monthly subtotals
            month_subs = {}
            for m in months:
                m_recs = [r for r in ch_recs if r.get("month") == f"{m}월"]
                month_subs[str(m)] = {
                    "discovered": len(m_recs),
                    "d_discovered": sum(1 for r in m_recs if r.get("grade_before") == "D"),
                    "improved": sum(1 for r in m_recs if r.get("completion") == "완료"),
                    "d_improved": sum(1 for r in m_recs if r.get("grade_before") == "D" and r.get("completion") == "완료"),
                }

            # Quarter total
            q_discovered = len(ch_recs)
            q_improved = sum(1 for r in ch_recs if r.get("completion") == "완료")
            q_d_disc = sum(1 for r in ch_recs if r.get("grade_before") == "D")
            q_d_imp = sum(1 for r in ch_recs if r.get("grade_before") == "D" and r.get("completion") == "완료")
            q_rate = round(q_improved / q_discovered, 4) if q_discovered > 0 else 0

            site_data[ch_name] = {
                "prev_discovered": prev_discovered,
                "prev_improved": prev_improved,
                "prev_rate": prev_rate,
                "weeks": weeks_data,
                "month_subs": month_subs,
                "quarter_discovered": q_discovered,
                "quarter_improved": q_improved,
                "quarter_d_discovered": q_d_disc,
                "quarter_d_improved": q_d_imp,
                "quarter_rate": q_rate,
            }

        # 합계
        total = {"prev_discovered": 0, "prev_improved": 0, "weeks": {}, "month_subs": {}, "quarter_discovered": 0, "quarter_improved": 0, "quarter_d_discovered": 0, "quarter_d_improved": 0}
        for ch in CHANNEL_ORDER:
            d = site_data[ch]
            total["prev_discovered"] += d["prev_discovered"]
            total["prev_improved"] += d["prev_improved"]
            total["quarter_discovered"] += d["quarter_discovered"]
            total["quarter_improved"] += d["quarter_improved"]
            total["quarter_d_discovered"] += d["quarter_d_discovered"]
            total["quarter_d_improved"] += d["quarter_d_improved"]
            for wk, wv in d["weeks"].items():
                if wk not in total["weeks"]:
                    total["weeks"][wk] = {"discovered": 0, "d_discovered": 0, "improved": 0, "d_improved": 0}
                for k2 in ("discovered", "d_discovered", "improved", "d_improved"):
                    total["weeks"][wk][k2] += wv[k2]
            for mk, mv in d["month_subs"].items():
                if mk not in total["month_subs"]:
                    total["month_subs"][mk] = {"discovered": 0, "d_discovered": 0, "improved": 0, "d_improved": 0}
                for k2 in ("discovered", "d_discovered", "improved", "d_improved"):
                    total["month_subs"][mk][k2] += mv[k2]
        total["prev_rate"] = round(total["prev_improved"] / total["prev_discovered"], 4) if total["prev_discovered"] > 0 else 0
        total["quarter_rate"] = round(total["quarter_improved"] / total["quarter_discovered"], 4) if total["quarter_discovered"] > 0 else 0
        site_data["합계"] = total

        result[site_name] = site_data

    return {
        "year": year,
        "quarter": quarter,
        "months": months,
        "channel_order": CHANNEL_ORDER,
        "sites": result,
    }


@app.get("/api/weekly/quarter")
async def weekly_quarter(request: Request, year: str = "2026", quarter: int = 1):
    """분기 전체 실시간 집계"""
    verify_token(request)
    records = load_data()
    stats = compute_quarter_stats(records, year, quarter)
    return stats


@app.post("/api/weekly/save")
async def weekly_save(request: Request):
    """관리자가 저장 - 분기 스냅샷 확정"""
    verify_admin(request)
    body = await request.json()
    year = int(body.get("year", 2026))
    quarter = int(body.get("quarter", 1))
    current_month = int(body.get("current_month", 1))
    current_week = int(body.get("current_week", 1))

    snapshot_id = f"{year}-Q{quarter}-{current_month}월{current_week}주"

    if DATABASE_URL:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id FROM weekly_snapshots WHERE id = %s", (snapshot_id,))
        if cur.fetchone():
            cur.close()
            release_db(conn)
            raise HTTPException(status_code=409, detail=f"{snapshot_id}는 이미 저장되었습니다.")
        cur.close()
        release_db(conn)

    records = load_data()
    stats = compute_quarter_stats(records, str(year), quarter)
    stats["current_month"] = current_month
    stats["current_week"] = current_week

    if DATABASE_URL:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO weekly_snapshots (id, year, quarter, month, week, saved_at, data) VALUES (%s, %s, %s, %s, %s, %s, %s)",
            (snapshot_id, year, quarter, current_month, current_week,
             datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
             json.dumps(stats, ensure_ascii=False))
        )
        conn.commit()
        cur.close()
        release_db(conn)

    return {"message": f"{snapshot_id} 저장 완료", "id": snapshot_id}


@app.get("/api/weekly/list")
async def weekly_list(request: Request, year: int = 2026, quarter: int = 0):
    """저장된 스냅샷 목록 조회"""
    verify_token(request)
    if not DATABASE_URL:
        return {"snapshots": []}
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if quarter > 0:
        cur.execute("SELECT id, year, quarter, month, week, saved_at FROM weekly_snapshots WHERE year = %s AND quarter = %s ORDER BY month, week", (year, quarter))
    else:
        cur.execute("SELECT id, year, quarter, month, week, saved_at FROM weekly_snapshots WHERE year = %s ORDER BY quarter, month, week", (year,))
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    release_db(conn)
    return {"snapshots": rows}


@app.get("/api/weekly/get")
async def weekly_get(request: Request, id: str = ""):
    """저장된 스냅샷 조회"""
    verify_token(request)
    if not DATABASE_URL or not id:
        return {"snapshot": None}
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM weekly_snapshots WHERE id = %s", (id,))
    row = cur.fetchone()
    cur.close()
    release_db(conn)
    if not row:
        raise HTTPException(status_code=404, detail="스냅샷을 찾을 수 없습니다.")
    row = dict(row)
    if isinstance(row["data"], str):
        row["data"] = json.loads(row["data"])
    return {"snapshot": row}


@app.delete("/api/weekly/delete")
async def weekly_delete(request: Request, id: str = ""):
    """관리자가 저장된 스냅샷 삭제"""
    verify_admin(request)
    if not DATABASE_URL or not id:
        raise HTTPException(status_code=400, detail="ID가 필요합니다.")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM weekly_snapshots WHERE id = %s", (id,))
    conn.commit()
    cur.close()
    release_db(conn)
    return {"message": f"{id} 삭제 완료"}


@app.get("/api/health")
async def health_check():
    db_url = os.environ.get("DATABASE_URL")
    all_db_vars = {k: v[:20] + "..." for k, v in os.environ.items() if "DATABASE" in k.upper() or "POSTGRES" in k.upper() or "PG" in k.upper()}
    result = {"database_url_set": bool(db_url), "db_connected": False, "table_exists": False, "env_vars": all_db_vars}
    if db_url:
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM records")
            count = cur.fetchone()[0]
            result["db_connected"] = True
            result["table_exists"] = True
            result["record_count"] = count
            cur.close()
            release_db(conn)
        except Exception as e:
            result["error"] = str(e)
    return result


# --- Excel Download ---
@app.get("/api/download-excel")
async def download_excel(request: Request, channel: str = "전체"):
    verify_token(request)
    records = load_data()
    if channel != "전체":
        records = [r for r in records if r.get("channel") == channel]
    if not records:
        raise HTTPException(status_code=404, detail=f"[{channel}] 데이터가 없습니다.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "위험요소"
    headers = ["No", "월", "부서", "담당자", "일시", "장소", "위험요소 내용", "공정", "재해유형",
               "가능성(전)", "중대성(전)", "위험도(전)", "등급(전)",
               "개선필요사항", "개선대책", "개선부서", "계획일", "완료일",
               "가능성(후)", "중대성(후)", "위험도(후)", "등급(후)",
               "완료여부", "비고", "추적관리자", "주차", "채널"]
    ws.append(headers)
    for r in records:
        ws.append([
            r.get("no", ""), r.get("month", ""), r.get("department", ""), r.get("person", ""),
            r.get("date", ""), r.get("location", ""), r.get("content_full", r.get("content", "")),
            r.get("process", ""), r.get("disaster_type", ""),
            r.get("likelihood_before", ""), r.get("severity_before", ""),
            r.get("risk_before", ""), r.get("grade_before", ""),
            r.get("improvement_needed", ""), r.get("improvement_plan", ""),
            r.get("improve_dept", ""), r.get("planned_date", ""), r.get("actual_date", ""),
            r.get("likelihood_after", ""), r.get("severity_after", ""),
            r.get("risk_after", ""), r.get("grade_after", ""),
            r.get("completion", ""), r.get("note", ""),
            r.get("tracking_manager", ""), r.get("week", ""), r.get("channel", ""),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"위험요소_{channel}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# --- Executive Comments & Notifications ---
ROLE_LABELS = {
    "1팀장": "환경안전1팀 팀장",
    "2팀장": "환경안전2팀 팀장",
    "본부장": "생산기술본부장",
    "부문장": "SCM 부문장",
    "대표이사": "대표이사",
}

COMMENTS_FILE = os.path.join(UPLOAD_DIR, "executive_comments.json")
NOTIFICATIONS_FILE = os.path.join(UPLOAD_DIR, "executive_notifications.json")


def get_week_key(dt=None):
    if dt is None:
        dt = datetime.now()
    iso = dt.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def format_week_label(week_key: str) -> str:
    parts = week_key.split("-W")
    if len(parts) != 2:
        return week_key
    year = int(parts[0])
    week = int(parts[1])
    jan4 = date(year, 1, 4)
    start = jan4 - timedelta(days=jan4.weekday()) + timedelta(weeks=week - 1)
    month = start.month
    first_of_month = date(start.year, month, 1)
    week_of_month = (start.day + first_of_month.weekday()) // 7 + 1
    return f"{year}년 {month}월 {week_of_month}주차"


def load_comments(week_key=None):
    if DATABASE_URL:
        conn = get_db()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if week_key:
            cur.execute("SELECT * FROM executive_comments WHERE week_key=%s ORDER BY created_at DESC", (week_key,))
        else:
            cur.execute("SELECT * FROM executive_comments ORDER BY created_at DESC")
        rows = cur.fetchall()
        cur.close()
        release_db(conn)
        return [dict(r) for r in rows]
    else:
        if os.path.exists(COMMENTS_FILE):
            with open(COMMENTS_FILE, "r", encoding="utf-8") as f:
                comments = json.load(f)
            if week_key:
                return [c for c in comments if c.get("week_key") == week_key]
            return comments
        return []


def get_comment_weeks():
    if DATABASE_URL:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT week_key FROM executive_comments ORDER BY week_key DESC")
        rows = cur.fetchall()
        cur.close()
        release_db(conn)
        return [r[0] for r in rows]
    else:
        if os.path.exists(COMMENTS_FILE):
            with open(COMMENTS_FILE, "r", encoding="utf-8") as f:
                comments = json.load(f)
            weeks = sorted(set(c.get("week_key", "") for c in comments), reverse=True)
            return [w for w in weeks if w]
        return []


def save_comments(comments):
    if not DATABASE_URL:
        with open(COMMENTS_FILE, "w", encoding="utf-8") as f:
            json.dump(comments, f, ensure_ascii=False)


def load_notifications():
    if DATABASE_URL:
        conn = get_db()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM executive_notifications WHERE is_read = FALSE ORDER BY created_at DESC")
        rows = cur.fetchall()
        cur.close()
        release_db(conn)
        return [dict(r) for r in rows]
    else:
        if os.path.exists(NOTIFICATIONS_FILE):
            with open(NOTIFICATIONS_FILE, "r", encoding="utf-8") as f:
                return [n for n in json.load(f) if not n.get("is_read")]
        return []


def save_notifications(notifications):
    if not DATABASE_URL:
        with open(NOTIFICATIONS_FILE, "w", encoding="utf-8") as f:
            json.dump(notifications, f, ensure_ascii=False)


@app.get("/api/comments")
async def get_comments(week: str = None):
    wk = week or get_week_key()
    return {"comments": load_comments(wk), "week_key": wk}


@app.get("/api/comment-weeks")
async def get_comment_weeks_api():
    weeks = get_comment_weeks()
    current = get_week_key()
    if current not in weeks:
        weeks.insert(0, current)
    result = []
    for w in weeks:
        parts = w.split("-W")
        if len(parts) != 2:
            continue
        year = int(parts[0])
        iso_week = int(parts[1])
        jan4 = date(year, 1, 4)
        start = jan4 - timedelta(days=jan4.weekday()) + timedelta(weeks=iso_week - 1)
        month = start.month
        first_of_month = date(start.year, month, 1)
        week_of_month = (start.day + first_of_month.weekday()) // 7 + 1
        month_key = f"{year}-{month:02d}"
        month_label = f"{year}년 {month}월"
        result.append({
            "key": w,
            "month_key": month_key,
            "month_label": month_label,
            "week_of_month": week_of_month,
        })
    return {"weeks": result}


@app.post("/api/comments")
async def create_comment(request: Request):
    body = await request.json()
    role = body.get("role", "")
    content = body.get("content", "").strip()
    if role not in ROLE_LABELS:
        raise HTTPException(status_code=400, detail="유효하지 않은 역할입니다.")
    if not content:
        raise HTTPException(status_code=400, detail="코멘트 내용을 입력해주세요.")

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    week_key = get_week_key()
    comment_id = uuid.uuid4().hex
    notif_id = uuid.uuid4().hex
    label = ROLE_LABELS[role]
    week_label = format_week_label(week_key)
    notif_msg = f"{label}님이 코멘트를 달았습니다 ({week_label})"

    comment = {"id": comment_id, "role": role, "content": content, "week_key": week_key, "created_at": now, "updated_at": None}
    notification = {"id": notif_id, "comment_id": comment_id, "message": notif_msg, "created_at": now, "is_read": False}

    if DATABASE_URL:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("INSERT INTO executive_comments (id, role, content, week_key, created_at) VALUES (%s,%s,%s,%s,%s)",
                     (comment_id, role, content, week_key, now))
        cur.execute("INSERT INTO executive_notifications (id, comment_id, message, created_at) VALUES (%s,%s,%s,%s)",
                     (notif_id, comment_id, notif_msg, now))
        conn.commit()
        cur.close()
        release_db(conn)
    else:
        comments = load_comments()
        comments.insert(0, comment)
        save_comments(comments)
        notifs = []
        if os.path.exists(NOTIFICATIONS_FILE):
            with open(NOTIFICATIONS_FILE, "r", encoding="utf-8") as f:
                notifs = json.load(f)
        notifs.insert(0, notification)
        save_notifications(notifs)

    return {"message": "코멘트 등록 완료", "comment": comment, "notification": notification}


@app.put("/api/comments/{comment_id}")
async def update_comment(comment_id: str, request: Request):
    body = await request.json()
    content = body.get("content", "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="코멘트 내용을 입력해주세요.")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if DATABASE_URL:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("UPDATE executive_comments SET content=%s, updated_at=%s WHERE id=%s", (content, now, comment_id))
        conn.commit()
        cur.close()
        release_db(conn)
    else:
        comments = load_comments()
        for c in comments:
            if c["id"] == comment_id:
                c["content"] = content
                c["updated_at"] = now
                break
        save_comments(comments)

    return {"message": "코멘트 수정 완료"}


@app.delete("/api/comments/{comment_id}")
async def delete_comment(comment_id: str):
    if DATABASE_URL:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM executive_notifications WHERE comment_id=%s", (comment_id,))
        cur.execute("DELETE FROM executive_comments WHERE id=%s", (comment_id,))
        conn.commit()
        cur.close()
        release_db(conn)
    else:
        comments = load_comments()
        comments = [c for c in comments if c["id"] != comment_id]
        save_comments(comments)
        notifs = []
        if os.path.exists(NOTIFICATIONS_FILE):
            with open(NOTIFICATIONS_FILE, "r", encoding="utf-8") as f:
                notifs = json.load(f)
        notifs = [n for n in notifs if n.get("comment_id") != comment_id]
        save_notifications(notifs)

    return {"message": "코멘트 삭제 완료"}


@app.get("/api/notifications")
async def get_notifications(current_week: bool = False):
    notifs = load_notifications()
    if current_week:
        wk = get_week_key()
        notifs = [n for n in notifs if n.get("created_at", "")[:10] >= _week_start(wk)]
    return {"notifications": notifs}


def _week_start(week_key: str) -> str:
    parts = week_key.split("-W")
    if len(parts) != 2:
        return ""
    year = int(parts[0])
    week = int(parts[1])
    jan4 = date(year, 1, 4)
    start = jan4 - timedelta(days=jan4.weekday()) + timedelta(weeks=week - 1)
    return start.isoformat()


@app.post("/api/notifications/{notification_id}/dismiss")
async def dismiss_notification(notification_id: str):
    if DATABASE_URL:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("UPDATE executive_notifications SET is_read=TRUE WHERE id=%s", (notification_id,))
        conn.commit()
        cur.close()
        release_db(conn)
    else:
        notifs = []
        if os.path.exists(NOTIFICATIONS_FILE):
            with open(NOTIFICATIONS_FILE, "r", encoding="utf-8") as f:
                notifs = json.load(f)
        for n in notifs:
            if n["id"] == notification_id:
                n["is_read"] = True
                break
        save_notifications(notifs)

    return {"message": "알림 확인 완료"}


@app.delete("/api/notifications/{notification_id}")
async def delete_notification(notification_id: str):
    if DATABASE_URL:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM executive_notifications WHERE id=%s", (notification_id,))
        conn.commit()
        cur.close()
        release_db(conn)
    else:
        notifs = []
        if os.path.exists(NOTIFICATIONS_FILE):
            with open(NOTIFICATIONS_FILE, "r", encoding="utf-8") as f:
                notifs = json.load(f)
        notifs = [n for n in notifs if n["id"] != notification_id]
        save_notifications(notifs)

    return {"message": "알림 삭제 완료"}


# --- Static Files ---
app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/")
async def root():
    return FileResponse("static/index.html")


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
